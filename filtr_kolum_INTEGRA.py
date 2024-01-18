import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def get_latest_file():
    files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('Modyfikowany')]
    if not files:
        return None
    latest_file = max(files, key=os.path.getctime)
    return latest_file

def auto_adjust_columns_width(file_path):
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    workbook.save(file_path)

def main():
    file_path = get_latest_file()
    if not file_path:
        print("Nie znaleziono odpowiedniego pliku Excel.")
        return

    with open("kolumny.txt", "r", encoding='utf-8') as file:
        columns = [line.strip() for line in file]

    df = pd.read_excel(file_path)

    missing_columns = [col for col in columns if col not in df.columns]
    if missing_columns:
        with open("brak_kolumn.txt", "w", encoding='utf-8') as f:
            for col in missing_columns:
                f.write(col + "\n")
        print(f"Utworzono plik 'brak_kolumn.txt' z brakującymi kolumnami: {', '.join(missing_columns)}")

    df_selected = df[columns]

    new_file_path = f"Modyfikowany_{os.path.basename(file_path)}"
    df_selected.to_excel(new_file_path, index=False)
    auto_adjust_columns_width(new_file_path)

    print(f"Plik został zmodyfikowany i zapisany jako {new_file_path}.")

if __name__ == "__main__":
    main()
